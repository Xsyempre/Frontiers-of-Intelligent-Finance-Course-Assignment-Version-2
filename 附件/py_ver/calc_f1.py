# -*- coding: utf-8 -*-
"""
F1 Score 计算脚本
计算三模型在 5% 样本上的 F1 Score

使用方法：
  python calc_f1.py
"""

import sys
import io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BENCHMARK_FILE = "../benchmark_G02.xlsx"
DEEPSEEK_FILE = "../deepseek_ver.xlsx"
MINIMAX_FILE = "../minimax_ver.xlsx"
MIMO_FILE = "../mimo_ver.xlsx"


def load_annotations(filename, col_offset):
    """从 Excel 文件加载标注结果"""
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    annotations = {}
    for r in range(4, ws.max_row + 1):
        seq = ws.cell(row=r, column=1).value  # SampleSeq
        if seq is None:
            continue
        seq = int(seq)

        ann = {}
        # 根据文件结构读取标注列
        # benchmark: col 24=ann_related, col 26=ann_fin_flag, col 28=third_party_flag
        # model files: same structure
        ann['ann_related'] = ws.cell(row=r, column=24).value
        ann['ann_fin_flag'] = ws.cell(row=r, column=26).value
        ann['third_party_flag'] = ws.cell(row=r, column=28).value
        annotations[seq] = ann

    return annotations


def calc_f1(human, model, field, subset=None):
    """计算单个字段的 F1 Score"""
    tp = fp = fn = tn = 0

    for seq in human:
        if seq not in model:
            continue

        h_val = human[seq].get(field)
        m_val = model[seq].get(field)

        # 处理 null 值
        if h_val is None or str(h_val).lower() in ('null', ''):
            continue
        if m_val is None or str(m_val).lower() in ('null', ''):
            continue

        h = int(h_val)
        m = int(m_val)

        # 子集过滤
        if subset:
            skip = False
            for sf, sv in subset.items():
                sv_val = human[seq].get(sf)
                if sv_val is None or str(sv_val).lower() in ('null', ''):
                    skip = True
                    break
                if int(sv_val) != sv:
                    skip = True
                    break
            if skip:
                continue

        if h == 1 and m == 1: tp += 1
        elif h == 0 and m == 1: fp += 1
        elif h == 1 and m == 0: fn += 1
        elif h == 0 and m == 0: tn += 1

    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

    return f1, tp, fp, fn, tn


def main():
    print("加载数据...")
    human = load_annotations(BENCHMARK_FILE, 0)
    deepseek = load_annotations(DEEPSEEK_FILE, 0)
    minimax = load_annotations(MINIMAX_FILE, 0)
    mimo = load_annotations(MIMO_FILE, 0)

    print(f"人工标注: {len(human)} 行")
    print(f"DeepSeek: {len(deepseek)} 行")
    print(f"MiniMax: {len(minimax)} 行")
    print(f"Mimo: {len(mimo)} 行")
    print()

    models = [("DeepSeek", deepseek), ("MiniMax", minimax), ("Mimo", mimo)]

    print(f"{'模型':<12} {'ann_related':>12} {'ann_fin_flag':>12} {'third_party':>12} {'平均':>8}")
    print("-" * 60)

    for name, model in models:
        f1_ar, *_ = calc_f1(human, model, 'ann_related')
        f1_af, *_ = calc_f1(human, model, 'ann_fin_flag', subset={'ann_related': 1})
        f1_tp, *_ = calc_f1(human, model, 'third_party_flag', subset={'ann_related': 1, 'ann_fin_flag': 1})

        valid = [f for f in [f1_ar, f1_af, f1_tp] if f > 0]
        avg = sum(valid) / len(valid) if valid else 0

        print(f"{name:<12} {f1_ar:>12.4f} {f1_af:>12.4f} {f1_tp:>12.4f} {avg:>8.4f}")


if __name__ == "__main__":
    main()
