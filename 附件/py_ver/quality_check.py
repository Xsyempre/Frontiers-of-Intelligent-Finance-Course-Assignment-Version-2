# -*- coding: utf-8 -*-
"""
标注质量检查与修正脚本（并行版）
三层检查：结构规则 → 抽样语义 → 边界案例

使用方法：
  python quality_check.py --phase structure
  python quality_check.py --phase sample
  python quality_check.py --phase boundary
  python quality_check.py --phase all
"""

import sys
import io
import os
import json
import re
import time
import random
import shutil
import argparse
import requests
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from config import BASE_URL, API_KEY, MODEL, MAX_WORKERS, PROMPT_FILE

INPUT_FILE = "STK_labeled_夏思远_2023110537_G02.xlsx"
BACKUP_FILE = "STK_labeled_backup.xlsx"
LOG_FILE = "quality_check_log.md"

VALID_ELEMENTS = {"资产", "负债", "所有者权益", "收入", "费用", "利润"}


def load_prompt(prompt_file):
    """从外部文件加载 System Prompt"""
    with open(prompt_file, 'r', encoding='utf-8') as f:
        return f.read().strip()


def safe_json_parse(s):
    if s is None or s == "" or s == "null":
        return None
    if isinstance(s, (list, dict)):
        return s
    try:
        return json.loads(str(s))
    except:
        return None


def is_null_value(v):
    if v is None:
        return True
    if isinstance(v, str) and (v.strip() == "" or v.strip().lower() == "null"):
        return True
    return False


def write_log(msg, mode="a"):
    with open(LOG_FILE, mode, encoding="utf-8") as f:
        f.write(msg + "\n")


def call_api(system_prompt, activity_text, max_retries=2):
    """调用 API 标注单行"""
    url = BASE_URL.rstrip('/') + '/chat/completions'
    headers = {'Authorization': f'Bearer {API_KEY}', 'Content-Type': 'application/json'}

    for attempt in range(max_retries):
        try:
            payload = {
                'model': MODEL,
                'messages': [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"请标注：\n\n{activity_text}"}
                ],
                'temperature': 0.1,
                'max_tokens': 2048
            }
            resp = requests.post(url, headers=headers, json=payload, timeout=60)
            data = resp.json()
            content = data['choices'][0]['message'].get('content', '')
            content = re.sub(r'<think>[\s\S]*?</think>', '', content, flags=re.DOTALL).strip()
            json_match = re.search(r'\{[\s\S]*\}', content)
            if json_match:
                return json.loads(json_match.group(0))
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(1)
    return None


def call_api_parallel(system_prompt, rows_data, desc="API调用"):
    """并行调用 API"""
    results = {}
    done = 0
    total = len(rows_data)
    start = time.time()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(call_api, system_prompt, act): row for row, act in rows_data}
        for future in as_completed(futures):
            row = futures[future]
            try:
                results[row] = future.result()
            except:
                results[row] = None
            done += 1
            if done % 10 == 0 or done == total:
                elapsed = time.time() - start
                speed = done / elapsed if elapsed > 0 else 0
                print(f"  {desc}: {done}/{total} ({speed:.1f} 条/秒)")

    return results


# ============================================================
# 第一层：结构规则检查
# ============================================================

def phase_structure(wb, ws):
    print("=" * 50)
    print("  第一层：结构规则检查")
    print("=" * 50)

    violations = {f"T{i}": [] for i in range(1, 8)}
    fixes = 0
    total = ws.max_row - 3

    for r in range(4, ws.max_row + 1):
        ar = ws.cell(row=r, column=24).value
        ay = ws.cell(row=r, column=25).value
        af = ws.cell(row=r, column=26).value
        ai = ws.cell(row=r, column=27).value
        tp = ws.cell(row=r, column=28).value
        tl = ws.cell(row=r, column=29).value

        if is_null_value(ar):
            violations["T1"].append(r)
            continue

        ar = int(ar)

        if ar == 0:
            for col in [25, 26, 27, 28, 29]:
                v = ws.cell(row=r, column=col).value
                if not is_null_value(v):
                    violations["T2"].append(r)
                    ws.cell(row=r, column=col, value="null")
                    fixes += 1
                    break

        if ar == 1:
            if is_null_value(ay):
                violations["T3"].append(r)
            if not is_null_value(af):
                af = int(af)
                if af == 0:
                    if not is_null_value(ai):
                        violations["T4"].append(r)
                        ws.cell(row=r, column=27, value="null")
                        fixes += 1
                    if not is_null_value(tp):
                        ws.cell(row=r, column=28, value="null")
                        fixes += 1
                    if not is_null_value(tl):
                        ws.cell(row=r, column=29, value="null")
                        fixes += 1
                elif af == 1:
                    if is_null_value(ai):
                        violations["T5"].append(r)
                    if not is_null_value(tp):
                        tp = int(tp)
                        if tp == 0 and not is_null_value(tl):
                            violations["T6"].append(r)
                            ws.cell(row=r, column=29, value="null")
                            fixes += 1
                        elif tp == 1 and is_null_value(tl):
                            violations["T7"].append(r)

    print(f"\n总数据行: {total}")
    print(f"\n规则违反:")
    for t in range(1, 8):
        key = f"T{t}"
        c = len(violations[key])
        print(f"  {key}: {c} {'PASS' if c == 0 else 'WARN'}")
    print(f"\n自动修正: {fixes} 处")

    write_log("# 质量检查日志\n", mode="w")
    write_log(f"检查时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
    write_log("## 第一层：结构规则检查\n")
    for t in range(1, 8):
        write_log(f"- T{t}: {len(violations[f'T{t}'])} 行违反")
    write_log(f"- 自动修正: {fixes} 处\n")

    return violations


# ============================================================
# 第二层：抽样语义检查（并行）
# ============================================================

def phase_sample(wb, ws, system_prompt, sample_size=30):
    print("\n" + "=" * 50)
    print("  第二层：抽样语义检查（并行）")
    print("=" * 50)

    rows = []
    for r in range(4, ws.max_row + 1):
        activity = ws.cell(row=r, column=17).value
        if activity and str(activity).strip() and activity != "没有单位":
            rows.append(r)

    random.seed(42)
    sample_rows = random.sample(rows, min(sample_size, len(rows)))

    current_annotations = {}
    for r in sample_rows:
        current_annotations[r] = {
            "ann_related": ws.cell(row=r, column=24).value,
            "ann_fin_flag": ws.cell(row=r, column=26).value,
            "third_party_flag": ws.cell(row=r, column=28).value,
        }

    rows_data = [(r, str(ws.cell(row=r, column=17).value)[:500]) for r in sample_rows]
    api_results = call_api_parallel(system_prompt, rows_data, desc="抽样验证")

    differences = []
    for r in sample_rows:
        api = api_results.get(r)
        if api is None:
            continue
        cur = current_annotations[r]
        for key in ["ann_related", "ann_fin_flag", "third_party_flag"]:
            cv = cur[key]
            av = api.get(key)
            if is_null_value(cv): cv = None
            if is_null_value(av): av = None
            if str(cv) != str(av):
                differences.append({"row": r, "field": key, "current": cv, "api": av,
                                    "activity": str(ws.cell(row=r, column=17).value)[:60]})
                break

    print(f"\n结果: {len(sample_rows)} 行中 {len(differences)} 行有差异 ({len(differences)/len(sample_rows)*100:.1f}%)")

    write_log("## 第二层：抽样语义检查\n")
    write_log(f"- 抽样: {len(sample_rows)} 行")
    write_log(f"- 差异: {len(differences)} 行 ({len(differences)/len(sample_rows)*100:.1f}%)\n")
    if differences:
        write_log("| 行号 | 字段 | 当前 | API | Activity |")
        write_log("|------|------|------|-----|----------|")
        for d in differences:
            write_log(f"| {d['row']} | {d['field']} | {d['current']} | {d['api']} | {d['activity']}... |")
        write_log("")

    return differences


# ============================================================
# 第三层：边界案例检查（并行）
# ============================================================

def phase_boundary(wb, ws, system_prompt):
    print("\n" + "=" * 50)
    print("  第三层：边界案例检查（并行）")
    print("=" * 50)

    keywords = ["年报", "年度报告"]
    false_pos = []
    false_neg = []

    for r in range(4, ws.max_row + 1):
        activity = str(ws.cell(row=r, column=17).value or "")
        ar = ws.cell(row=r, column=24).value
        if is_null_value(ar):
            continue
        ar = int(ar)
        has_kw = any(kw in activity for kw in keywords)
        if ar == 1 and not has_kw:
            false_pos.append(r)
        elif ar == 0 and has_kw:
            false_neg.append(r)

    print(f"ann_related=1 无关键词: {len(false_pos)} 行")
    print(f"ann_related=0 有关键词: {len(false_neg)} 行")

    check_rows = false_pos[:20] + false_neg[:20]
    print(f"并行验证 {len(check_rows)} 行...")

    rows_data = [(r, str(ws.cell(row=r, column=17).value or "")[:500]) for r in check_rows]
    api_results = call_api_parallel(system_prompt, rows_data, desc="边界验证")

    corrections = []
    for r in check_rows:
        api = api_results.get(r)
        if api is None:
            continue
        cur_ar = int(ws.cell(row=r, column=24).value)
        api_ar = api.get("ann_related")
        if api_ar is not None and int(api_ar) != cur_ar:
            corrections.append({"row": r, "current": cur_ar, "api": int(api_ar),
                                "activity": str(ws.cell(row=r, column=17).value)[:60]})

    print(f"需修正: {len(corrections)} 行")

    for c in corrections:
        r = c["row"]
        new_ar = c["api"]
        ws.cell(row=r, column=24, value=new_ar)
        if new_ar == 0:
            for col in [25, 26, 27, 28, 29]:
                ws.cell(row=r, column=col, value="null")
        else:
            activity = str(ws.cell(row=r, column=17).value or "")[:500]
            full = call_api(system_prompt, activity)
            if full:
                ws.cell(row=r, column=25, value=json.dumps(full.get("ann_year"), ensure_ascii=False) if full.get("ann_year") else "null")
                ws.cell(row=r, column=26, value=full.get("ann_fin_flag") if full.get("ann_fin_flag") is not None else "null")
                ws.cell(row=r, column=27, value=json.dumps(full.get("ann_fin_info"), ensure_ascii=False) if full.get("ann_fin_info") else "null")
                ws.cell(row=r, column=28, value=full.get("third_party_flag") if full.get("third_party_flag") is not None else "null")
                ws.cell(row=r, column=29, value=json.dumps(full.get("third_party_list"), ensure_ascii=False) if full.get("third_party_list") else "null")

    write_log("## 第三层：边界案例检查\n")
    write_log(f"- ann_related=1 无关键词: {len(false_pos)} 行")
    write_log(f"- ann_related=0 有关键词: {len(false_neg)} 行")
    write_log(f"- 需修正: {len(corrections)} 行\n")
    if corrections:
        write_log("| 行号 | 原值 | 修正 | Activity |")
        write_log("|------|------|------|----------|")
        for c in corrections:
            write_log(f"| {c['row']} | {c['current']} | {c['api']} | {c['activity']}... |")
        write_log("")

    return corrections


# ============================================================
# 主流程
# ============================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--phase", default="all", choices=["structure", "sample", "boundary", "all"])
    args = parser.parse_args()

    system_prompt = load_prompt(PROMPT_FILE)

    print("=" * 50)
    print("  标注质量检查（并行版）")
    print("=" * 50)
    print(f"输入: {INPUT_FILE}")
    print(f"模型: {MODEL} ({MAX_WORKERS}线程并行)")
    print(f"Prompt: {PROMPT_FILE}")
    print()

    if not os.path.exists(BACKUP_FILE):
        shutil.copy2(INPUT_FILE, BACKUP_FILE)
        print(f"已备份: {BACKUP_FILE}")

    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    print(f"总行数: {ws.max_row - 3}\n")

    start = time.time()

    if args.phase in ["structure", "all"]:
        phase_structure(wb, ws)
        wb.save(INPUT_FILE)
        print("  已保存\n")

    if args.phase in ["sample", "all"]:
        phase_sample(wb, ws, system_prompt)
        wb.save(INPUT_FILE)
        print("  已保存\n")

    if args.phase in ["boundary", "all"]:
        phase_boundary(wb, ws, system_prompt)
        wb.save(INPUT_FILE)
        print("  已保存\n")

    elapsed = time.time() - start
    print("=" * 50)
    print(f"  完成！耗时: {elapsed/60:.1f} 分钟")
    print(f"  日志: {LOG_FILE}")
    print("=" * 50)


if __name__ == "__main__":
    main()
