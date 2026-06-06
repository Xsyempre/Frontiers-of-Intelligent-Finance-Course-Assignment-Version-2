# -*- coding: utf-8 -*-
"""
API 连通性测试脚本
测试前 5 行数据，验证 API 配置和 Prompt 是否正确

使用方法：
  python test_api.py
"""

import sys
import io
import json
import re
import requests
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from config import BASE_URL, API_KEY, MODEL, INPUT_FILE, PROMPT_FILE


def load_prompt(prompt_file):
    with open(prompt_file, 'r', encoding='utf-8') as f:
        return f.read().strip()


def call_api(system_prompt, activity_text):
    url = BASE_URL.rstrip('/') + '/chat/completions'
    headers = {'Authorization': f'Bearer {API_KEY}', 'Content-Type': 'application/json'}
    payload = {
        'model': MODEL,
        'messages': [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"请对以下违规公告进行标注：\n\n{activity_text}"}
        ],
        'temperature': 0.1,
        'max_tokens': 4096,
        'response_format': {'type': 'json_object'}
    }
    resp = requests.post(url, headers=headers, json=payload, timeout=120)
    data = resp.json()
    content = data['choices'][0]['message'].get('content', '')
    content = re.sub(r'<think>[\s\S]*?</think>', '', content, flags=re.DOTALL).strip()
    if content:
        return json.loads(content)
    return None


def main():
    print(f"模型: {MODEL}")
    print(f"API: {BASE_URL}")
    print(f"Prompt: {PROMPT_FILE}")
    print()

    system_prompt = load_prompt(PROMPT_FILE)
    print(f"Prompt 长度: {len(system_prompt)} 字符")
    print()

    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    # 测试前 5 行有效数据
    tested = 0
    for r in range(4, ws.max_row + 1):
        if tested >= 5:
            break

        activity = ws.cell(row=r, column=17).value
        if not activity or str(activity).strip() == "" or activity == "没有单位":
            continue

        tested += 1
        activity_text = str(activity)[:300]
        print(f"--- Row {r} ---")
        print(f"Activity: {activity_text}...")

        result = call_api(system_prompt, activity_text)
        if result:
            print(f"Result: {json.dumps(result, ensure_ascii=False)[:200]}")
        else:
            print("Result: API 调用失败")
        print()

    print(f"测试完成: {tested} 行")


if __name__ == "__main__":
    main()
