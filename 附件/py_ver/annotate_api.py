# -*- coding: utf-8 -*-
"""
全量财务舞弊标注 - API 版本（并行版）
使用大模型 API 进行智能标注，支持多线程并行调用

使用方法：
1. 修改 config.py 中的 API 配置
2. 确保 prompts/system_prompt.md 存在
3. 运行: python annotate_api.py [--input 文件] [--output 文件] [--limit N] [--workers N]
"""

import sys
import io
import json
import time
import re
import os
import requests
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from config import BASE_URL, API_KEY, MODEL, MAX_WORKERS, SAVE_INTERVAL, INPUT_FILE, OUTPUT_FILE, PROMPT_FILE


def load_prompt(prompt_file):
    """从外部文件加载 System Prompt"""
    with open(prompt_file, 'r', encoding='utf-8') as f:
        return f.read().strip()


def call_llm(system_prompt, activity_text, max_retries=3):
    """调用大模型 API 进行标注"""
    url = BASE_URL.rstrip('/') + '/chat/completions'
    headers = {'Authorization': f'Bearer {API_KEY}', 'Content-Type': 'application/json'}

    for attempt in range(max_retries):
        try:
            payload = {
                'model': MODEL,
                'messages': [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"请对以下违规公告进行标注：\n\n{activity_text}"}
                ],
                'temperature': 0.1,
                'max_tokens': 4096,
                'response_format': {'type': 'json_object'}
            }
            resp = requests.post(url, headers=headers, json=payload, timeout=120)
            data = resp.json()
            content = data['choices'][0]['message'].get('content', '')

            # 剥离推理模型的 <think> 标签
            content = re.sub(r'<think>[\s\S]*?</think>', '', content, flags=re.DOTALL).strip()

            if content:
                return json.loads(content)

        except json.JSONDecodeError:
            time.sleep(1)
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(2)

    return {
        "ann_related": 0, "ann_year": None,
        "ann_fin_flag": None, "ann_fin_info": None,
        "third_party_flag": None, "third_party_list": None,
    }


def validate_result(result):
    """验证并修正标注结果"""
    ar = result.get("ann_related", 0)
    if ar not in [0, 1]:
        ar = 0

    if ar == 0:
        return {"ann_related": 0, "ann_year": None, "ann_fin_flag": None,
                "ann_fin_info": None, "third_party_flag": None, "third_party_list": None}

    ay = result.get("ann_year")
    if ay and isinstance(ay, list):
        ay = sorted([int(y) for y in ay if 1990 <= int(y) <= 2030])
        if not ay:
            ay = None
    else:
        ay = None

    af = result.get("ann_fin_flag")
    if af not in [0, 1]:
        af = 0

    ai = result.get("ann_fin_info")
    if af == 1 and ai and isinstance(ai, list):
        valid_elements = {"资产", "负债", "所有者权益", "收入", "费用", "利润"}
        for item in ai:
            if "elements" in item:
                item["elements"] = [e for e in item["elements"] if e in valid_elements]
    elif af == 0:
        ai = None

    tp = result.get("third_party_flag")
    if tp not in [0, 1]:
        tp = 0

    tl = result.get("third_party_list")
    if tp == 1 and tl and isinstance(tl, list):
        valid_types = {"客户", "供应商", "银行/金融机构", "券商/保荐机构",
                       "会计师事务所", "评估机构", "自然人", "其他企业"}
        for item in tl:
            if "type" in item and item["type"] not in valid_types:
                item["type"] = "其他企业"
    elif tp == 0:
        tl = None

    return {"ann_related": ar, "ann_year": ay, "ann_fin_flag": af,
            "ann_fin_info": ai, "third_party_flag": tp, "third_party_list": tl}


def process_one(system_prompt, activity_text, src_row):
    """处理单行，返回 (src_row, result_dict)"""
    result = call_llm(system_prompt, str(activity_text))
    result = validate_result(result)
    return (src_row, result)


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=str, default=None, help="输入文件路径")
    parser.add_argument("--output", type=str, default=None, help="输出文件路径")
    parser.add_argument("--limit", type=int, default=0, help="只处理前 N 行（0=全量）")
    parser.add_argument("--workers", type=int, default=MAX_WORKERS, help="并发数")
    args = parser.parse_args()

    input_file = args.input or INPUT_FILE
    output_file = args.output or OUTPUT_FILE

    # 加载 Prompt（从外部文件）
    system_prompt = load_prompt(PROMPT_FILE)

    print(f"模型: {MODEL}")
    print(f"API: {BASE_URL}")
    print(f"Prompt: {PROMPT_FILE}")
    print(f"并发数: {args.workers}")
    if args.limit:
        print(f"限制: 前 {args.limit} 行")
    print()

    # 读取源数据
    print("正在读取数据...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    print(f"  总行数: {ws.max_row}, 列数: {ws.max_column}")

    # 创建输出工作簿
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active

    # 复制原始列头
    for col in range(1, ws.max_column + 1):
        out_ws.cell(row=1, column=col, value=ws.cell(row=1, column=col).value)

    # 添加标注列头
    new_cols = ["ann_related", "ann_year", "ann_fin_flag", "ann_fin_info",
                "third_party_flag", "third_party_list"]
    ann_start_col = ws.max_column + 1
    for i, h in enumerate(new_cols):
        out_ws.cell(row=1, column=ann_start_col + i, value=h)

    # 复制表头行（第2-3行）
    for r in range(2, 4):
        for col in range(1, ws.max_column + 1):
            out_ws.cell(row=r, column=col, value=ws.cell(row=r, column=col).value)
        for i in range(6):
            out_ws.cell(row=r, column=ann_start_col + i, value="")

    # 收集需要标注的行
    tasks = []
    empty_rows = []

    for src_row in range(4, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            out_ws.cell(row=src_row, column=col, value=ws.cell(row=src_row, column=col).value)

        activity = ws.cell(row=src_row, column=17).value
        if not activity or str(activity).strip() == "" or activity == "没有单位":
            empty_rows.append(src_row)
            for i in range(6):
                out_ws.cell(row=src_row, column=ann_start_col + i, value="null")
        else:
            tasks.append((src_row, str(activity)))

    print(f"  有效数据行: {len(tasks)}, 空行: {len(empty_rows)}")

    if args.limit:
        tasks = tasks[:args.limit]
        print(f"  限制处理: {len(tasks)} 行")

    # 并行标注
    total = len(tasks)
    done = 0
    errors = 0
    start_time = time.time()

    print(f"\n开始并行标注（{args.workers} 线程）...")

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {executor.submit(process_one, system_prompt, act, row): row for row, act in tasks}

        for future in as_completed(futures):
            src_row = futures[future]
            try:
                row_idx, result = future.result()

                out_ws.cell(row=row_idx, column=ann_start_col, value=result["ann_related"])
                out_ws.cell(row=row_idx, column=ann_start_col + 1,
                            value=json.dumps(result["ann_year"], ensure_ascii=False) if result["ann_year"] else "null")
                out_ws.cell(row=row_idx, column=ann_start_col + 2,
                            value=result["ann_fin_flag"] if result["ann_fin_flag"] is not None else "null")
                out_ws.cell(row=row_idx, column=ann_start_col + 3,
                            value=json.dumps(result["ann_fin_info"], ensure_ascii=False) if result["ann_fin_info"] else "null")
                out_ws.cell(row=row_idx, column=ann_start_col + 4,
                            value=result["third_party_flag"] if result["third_party_flag"] is not None else "null")
                out_ws.cell(row=row_idx, column=ann_start_col + 5,
                            value=json.dumps(result["third_party_list"], ensure_ascii=False) if result["third_party_list"] else "null")

            except Exception as e:
                errors += 1
                for i in range(6):
                    out_ws.cell(row=src_row, column=ann_start_col + i, value="null")

            done += 1
            if done % 10 == 0 or done == total:
                elapsed = time.time() - start_time
                speed = done / elapsed if elapsed > 0 else 0
                eta = (total - done) / speed if speed > 0 else 0
                print(f"  进度: {done}/{total} ({speed:.1f} 条/秒, 剩余 {eta/60:.0f} 分钟, 错误: {errors})")

            if done % SAVE_INTERVAL == 0:
                out_wb.save(output_file)

    out_wb.save(output_file)

    elapsed = time.time() - start_time
    print(f"\n=== 标注完成 ===")
    print(f"有效数据行: {total}")
    print(f"空行: {len(empty_rows)}")
    print(f"错误: {errors}")
    print(f"耗时: {elapsed/60:.1f} 分钟 ({elapsed:.0f} 秒)")
    print(f"速度: {total/elapsed:.1f} 条/秒")
    print(f"已保存: {output_file}")


if __name__ == "__main__":
    main()
